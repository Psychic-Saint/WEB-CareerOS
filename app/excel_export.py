"""
app/excel_export.py — WEB CareerOS™ Excel Export
Produces a WEB-branded .xlsx from the job queue.
Called by the Streamlit dashboard's Export button.

Requires: openpyxl  (pip install openpyxl)
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ── WEB brand colours ────────────────────────────────────────────
C_DARK        = "FF010D0F"   # #010d0f  darkest bg
C_CARD        = "FF021A1E"   # #021a1e  card bg
C_CYAN        = "FF00E5FF"   # #00e5ff  primary cyan
C_CYAN2       = "FF00BCD4"   # #00bcd4  secondary cyan
C_MUTED       = "FF7ECDD8"   # #7ecdd8  muted text
C_WHITE       = "FFE0F7FA"   # #e0f7fa  near-white
C_ACCENT      = "FF00BFA5"   # #00bfa5  teal accent
C_GREEN       = "FF00C853"   # #00c853  submitted
C_AMBER       = "FFFFB300"   # #ffb300  escalated
C_RED         = "FFFF5252"   # #ff5252  rejected
C_PURPLE      = "FFE040FB"   # #e040fb  failed

# Status → fill colour
STATUS_FILLS = {
    "submitted":      C_GREEN,
    "escalated":      C_AMBER,
    "rejected":       C_RED,
    "approved":       C_ACCENT,
    "new":            "FF607D8B",
    "scored":         "FFFFD740",
    "failed":         C_PURPLE,
    "skipped":        "FF546E7A",
    "pending_review": C_MUTED,
    "drafted":        C_MUTED,
}

COLUMNS = [
    ("Source",         16),
    ("Title",          36),
    ("Company",        22),
    ("Location",       18),
    ("Remote",         9),
    ("Status",         18),
    ("Action Taken",   22),
    ("Fit Score",      11),
    ("AI Reasoning",   50),
    ("Posted",         13),
    ("Updated",        13),
    ("Apply URL",      40),
]


def _fmt_date(iso: str) -> str:
    if not iso:
        return ""
    try:
        return datetime.fromisoformat(iso.replace("Z", "+00:00")).strftime("%d %b %Y")
    except Exception:
        return iso[:10] if len(iso) >= 10 else iso


def _action_label(job: dict) -> str:
    s = job.get("status", "")
    n = (job.get("notes") or "").lower()
    if s == "submitted":
        return "Email Sent" if ("email_sent_to" in n or "platform=email" in n) else "Form Submitted"
    if s == "escalated":
        if "captcha" in n:     return "CAPTCHA Block"
        if "manual_only" in n: return "Manual Platform"
        if "fill_rate" in n:   return "Form Incomplete"
        return "Escalated — Manual"
    if s == "rejected": return "Poor Fit"
    if s == "approved": return "Queued"
    return s.replace("_", " ").title()


def _fill(hex_argb: str) -> "PatternFill":
    return PatternFill(fill_type="solid", fgColor=hex_argb)


def _font(bold=False, color=C_WHITE, size=10) -> "Font":
    return Font(bold=bold, color=color, size=size, name="Segoe UI")


def _border() -> "Border":
    side = Side(style="thin", color="FF00E5FF22")
    return Border(left=side, right=side, top=side, bottom=side)


def build_excel_bytes(jobs: list[dict[str, Any]]) -> bytes:
    """
    Build a WEB-branded Excel workbook from a list of job dicts.
    Returns raw bytes suitable for st.download_button.
    Falls back to a minimal CSV-in-xlsx if openpyxl is missing.
    """
    if not _HAS_OPENPYXL:
        # Graceful fallback: TSV inside a byte wrapper that Excel can open
        lines = ["\t".join(h for h, _ in COLUMNS)]
        for j in jobs:
            lines.append("\t".join([
                j.get("source",""), j.get("title",""), j.get("company",""),
                j.get("location",""), "Yes" if j.get("is_remote") else "No",
                j.get("status",""), _action_label(j),
                str(j.get("fit_score") or ""),
                (j.get("fit_reasoning") or "")[:100],
                _fmt_date(j.get("posted_at","")),
                _fmt_date(j.get("updated_at","")),
                j.get("apply_url",""),
            ]))
        return "\n".join(lines).encode("utf-8")

    wb = openpyxl.Workbook()

    # ── Sheet 1: All Jobs ─────────────────────────────────────────
    ws = wb.active
    ws.title = "All Jobs"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "00E5FF"

    # Title row
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value  = "WEB CareerOS™  ·  Job Application Pipeline"
    title_cell.font   = Font(bold=True, color=C_CYAN, size=14, name="Segoe UI")
    title_cell.fill   = _fill(C_DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sub-title row
    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:L2")
    sub_cell = ws["A2"]
    sub_cell.value  = f"Exported {datetime.now().strftime('%d %b %Y  %H:%M')}  ·  {len(jobs)} jobs"
    sub_cell.font   = Font(bold=False, color=C_MUTED, size=9, name="Segoe UI")
    sub_cell.fill   = _fill(C_DARK)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Header row (row 3)
    ws.row_dimensions[3].height = 22
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = _font(bold=True, color=C_DARK, size=10)
        cell.fill      = _fill(C_CYAN)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze panes (freeze rows 1-3 + column A)
    ws.freeze_panes = "B4"

    # Data rows
    for row_idx, job in enumerate(jobs, start=4):
        ws.row_dimensions[row_idx].height = 16

        status      = job.get("status", "")
        score       = job.get("fit_score")
        status_fill = STATUS_FILLS.get(status, "FF455A64")
        row_bg      = C_CARD if row_idx % 2 == 0 else C_DARK

        values = [
            job.get("source", ""),
            job.get("title", ""),
            job.get("company", ""),
            job.get("location", ""),
            "🌐 Remote" if job.get("is_remote") else "📍 On-site",
            status,
            _action_label(job),
            score if score is not None else "",
            (job.get("fit_reasoning") or "")[:200],
            _fmt_date(job.get("posted_at", "")),
            _fmt_date(job.get("updated_at", "")),
            job.get("apply_url", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = _font(color=C_WHITE, size=9)
            cell.fill      = _fill(row_bg)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border    = _border()

            # Colour the Status column
            if col_idx == 6:
                cell.fill  = _fill(status_fill + "55")   # semi-transparent
                cell.font  = Font(bold=True, color=status_fill, size=9, name="Segoe UI")

            # Colour score: green ≥70, amber ≥55, red <55
            if col_idx == 8 and score is not None:
                sc = C_GREEN if score >= 70 else (C_AMBER if score >= 55 else C_RED)
                cell.font  = Font(bold=True, color=sc, size=9, name="Segoe UI")

            # Make URL clickable
            if col_idx == 12 and value and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(color=C_CYAN2, underline="single", size=9, name="Segoe UI")

    # ── Auto-filter on header row ─────────────────────────────────
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLUMNS))}{max(3, len(jobs)+3)}"

    # ── Sheet 2: Summary ─────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "00BFA5"
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14

    def _summary_header(row, text):
        ws2.row_dimensions[row].height = 22
        c = ws2.cell(row=row, column=1, value=text)
        c.font      = Font(bold=True, color=C_CYAN, size=11, name="Segoe UI")
        c.fill      = _fill(C_DARK)
        ws2.merge_cells(f"A{row}:B{row}")
        c.alignment = Alignment(horizontal="left", vertical="center")

    def _summary_row(row, label, value, value_color=C_WHITE):
        ws2.row_dimensions[row].height = 18
        lc = ws2.cell(row=row, column=1, value=label)
        lc.font      = _font(color=C_MUTED, size=10)
        lc.fill      = _fill(C_CARD if row % 2 == 0 else C_DARK)
        lc.alignment = Alignment(vertical="center")
        vc = ws2.cell(row=row, column=2, value=value)
        vc.font      = Font(bold=True, color=value_color, size=10, name="Segoe UI")
        vc.fill      = _fill(C_CARD if row % 2 == 0 else C_DARK)
        vc.alignment = Alignment(horizontal="center", vertical="center")

    from collections import Counter
    status_counts = Counter(j.get("status","") for j in jobs)
    source_counts = Counter(j.get("source","") for j in jobs)

    _summary_header(1, "WEB CareerOS™  —  Pipeline Summary")
    _summary_row(2, "Export Date", datetime.now().strftime("%d %b %Y  %H:%M"))
    _summary_row(3, "Total Jobs Discovered", len(jobs))
    _summary_row(4, "Auto-Applied (submitted)", status_counts.get("submitted",0), C_GREEN)
    _summary_row(5, "Need Your Action (escalated)", status_counts.get("escalated",0), C_AMBER)
    _summary_row(6, "Poor Fit (rejected)", status_counts.get("rejected",0), C_RED)
    _summary_row(7, "Queued (approved)", status_counts.get("approved",0), C_ACCENT)
    _summary_row(8, "New / Unscored", status_counts.get("new",0) + status_counts.get("scored",0))

    _summary_header(10, "Jobs by Source")
    for i, (src, cnt) in enumerate(sorted(source_counts.items(), key=lambda x: -x[1]), start=11):
        color = C_CYAN if i % 2 == 0 else C_MUTED
        _summary_row(i, src, cnt, color)

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
